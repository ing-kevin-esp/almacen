import pandas as pd
import pdb
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class Procesador:

    def __init__(self, inputExcel):
        self.writer = pd.ExcelWriter(inputExcel, engine='openpyxl', mode='a', if_sheet_exists='replace')

        self.matrizSheet = pd.read_excel(self.writer, sheet_name='matriz')
        self.actualizacionsheet = pd.read_excel(self.writer, sheet_name='actualizacion')

        self.hectorSheet = pd.read_excel(self.writer, sheet_name='hector')
        self.pedroSheet = pd.read_excel(self.writer, sheet_name='pedro')

        self.itemsNuevosSheet = None
        self.diferenciasSheet = None
        self.hectorResult = None
        self.pedroResult = None


    def search_new_items(self):
        df_merged = self.actualizacionsheet.merge(self.matrizSheet, how='left', left_on='iblitm', right_on='iblitm', indicator=True)
        filtered_items = df_merged.where(df_merged['_merge']=='left_only').dropna(thresh=2)

        self.itemsNuevosSheet = filtered_items[filtered_items.columns[:10]]

    def differences_with_matrix(self):
        """This function will join new information table with out matrix and list their quantities to find differences"""
        df_merged_all = self.actualizacionsheet.merge(self.matrizSheet, how='left', left_on='iblitm', right_on='iblitm', indicator=True)
        df_merged_inner = df_merged_all.where(df_merged_all['_merge']=='both').dropna(thresh=2)

        df_merged_inner_few_columns = df_merged_inner[['ibsrp1_y', 'iblitm' ,'descripcion_y', 'um_y', 'lipqoh', 'fisico', 'lilocn_y', 'c42_y','ClasABC_y', 'Empaque_y']]
        new_names = {
            'ibsrp1_y':'ibsrp1',
            'descripcion_y':'descripcion',
            'um_y': 'um',
            'lipqoh': 'sistema',
            'lilocn_y': 'lilocn',
            'c42_y': 'c42',
            'ClasABC_y':'ClasABC',
            'Empaque_y':'Empaque'
            }

        self.diferenciasSheet = df_merged_inner_few_columns.rename(columns=new_names)

    def hector(self):
        """This function will update hector table, something similar that differences_with_matrix does"""
        df_merged_all = self.diferenciasSheet.merge(self.hectorSheet, how='left', left_on='iblitm', right_on='iblitm', indicator=True)
        df_merged_inner = df_merged_all.where(df_merged_all['_merge']=='both').dropna(thresh=2)

        df_merged_inner_few_columns = df_merged_inner[['ibsrp1_y', 'iblitm' ,'descripcion_y', 'um_y',  'sistema','sistema_x','lilocn_y', 'c42_y', 'ClasABC_y', 'Empaque_y']]

        new_names = {
            'ibsrp1_y':'ibsrp1',
            'descripcion_y':'descripcion',
            'um_y': 'um',
            'sistema_y':'sist',
            'sistema_x': 'fisico',#aqui va duda, que columna es la que se tiene que tomar como "fisico ??"
            'lilocn_y': 'lilocn',
            'c42_y': 'c42',
            'ClasABC_y':'ClasABC',
            'Empaque_y':'Empaque'
            }

        self.hectorResult = df_merged_inner_few_columns.rename(columns=new_names)

    def pedro(self):
        """This function will update hector table, something similar that differences_with_matrix does"""
        df_merged_all = self.diferenciasSheet.merge(self.pedroSheet, how='left', left_on='iblitm', right_on='iblitm', indicator=True)
        df_merged_inner = df_merged_all.where(df_merged_all['_merge']=='both').dropna(thresh=2)

        df_merged_inner_few_columns = df_merged_inner[['ibsrp1_y', 'iblitm' ,'descripcion_y', 'um_y',  'sistema_y','sistema_x','lilocn_y', 'c42_y', 'ClasABC_y', 'Empaque_y']]
        new_names = {
            'ibsrp1_y':'ibsrp1',
            'descripcion_y':'descripcion',
            'um_y': 'um',
            'sist._y':'sist',
            'sist._x': 'fisico',#aqui va duda, que columna es la que se tiene que tomar como "fisico ??"
            'lilocn_y': 'lilocn',
            'c42_y': 'c42',
            'ClasABC_y':'ClasABC',
            'Empaque_y':'Empaque'
            }

        self.pedroResult = df_merged_inner_few_columns.rename(columns=new_names)

    def save(self):

        self.itemsNuevosSheet.to_excel(self.writer, sheet_name='items_nuevos', header=True, startrow=0, index=False)
        self.diferenciasSheet.to_excel(self.writer, sheet_name='matriz', header=True, startrow=0, index=False)
        #self.hectorResult.to_excel(self.writer, sheet_name='hector', header=True, startrow=0, index=False)
        #self.pedroResult.to_excel(self.writer, sheet_name='pedro', header=True, startrow=0, index=False)
        self.writer.close()


class Formateador:

    def __init__(self):
        self.work_book = None
        self.loadExcel()

    def loadExcel(self, file_name = 'FORMATO DE INVENTARIO.xlsx'):
        self.work_book = load_workbook(filename=file_name)

    def compareSistemaAndFisico(self, sheetName='matriz'):
        matrizSheet= self.work_book[sheetName]
        for row in matrizSheet.iter_rows(min_row=2, min_col=5, max_col=6):
            print(row)
            # Compare quantities between System and new
            if row[1].value < row[0].value:
                #rojo si es menor
                row[1].fill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')

            elif row[1].value > row[0].value:
                #verde si es mayor
                row[1].fill = PatternFill(start_color='8fce00',end_color='8fce00',fill_type='solid')

    def save(self):
        self.work_book.save("Nuevo formato de inventario.xlsx")



a = Procesador('FORMATO DE INVENTARIO.xlsx')
a.search_new_items()
a.differences_with_matrix()
#a.hector()
#a.pedro()
a.save()

b = Formateador()
b.loadExcel('FORMATO DE INVENTARIO.xlsx')
b.compareSistemaAndFisico()
b.save()

